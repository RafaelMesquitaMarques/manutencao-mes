"""Maintenance costs & budget tracking (Costs page).

Actuals come from the same sources as the KPI cost widgets: WOCost transactions
plus parts used (WO parts and approved intervention parts). Budgets are stored
per (year, month) in maintenance_budgets and edited from the Costs page — writes
are gated by the `costs` resource guard at router registration.
"""
import io
import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import date
from typing import List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import delete as sa_delete, distinct, extract, func, select, update as sa_update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import get_current_user
from app.core.plant_context import ERR_PLANT_NOT_AUTHORIZED, PlantContext, get_plant_context
from app.db.session import get_db
from app.models.models import (
    CostCenter, CostCenterBudget, CostCenterDepartment, CostTransactionType,
    Equipment, InterventionPart, LaborRecord, MachineIntervention, MaintenanceBudget,
    Plant, PurchaseOrder, SapCostLine, Supplier, User, WOCost, WOPart, WorkOrder,
)

router = APIRouter()

UNASSIGNED = "Unassigned"


# ─── Fiscal calendar (the SAP fiscal year runs Dec–Nov) ───────────────────────
#
# Every aggregate on this page is a 12-slot array. In calendar mode slot i is
# month i+1 of `year`; in fiscal mode (active when SAP lines exist for the year)
# slot 0 is December of year-1 and slots 1..11 are Jan–Nov of `year`. A months
# map lists the (calendar year, month) behind each slot so both the queries and
# the client can translate.

MonthsMap = List[tuple]  # 12 (year, month) pairs in display order


def _months_map(year: int, fiscal: bool) -> MonthsMap:
    if fiscal:
        return [(year - 1, 12)] + [(year, m) for m in range(1, 12)]
    return [(year, m) for m in range(1, 13)]


def _slot_index(mmap: MonthsMap) -> dict:
    """(calendar year, month) → 0-based slot in the months map."""
    return {ym: i for i, ym in enumerate(mmap)}


def _map_years(mmap: MonthsMap) -> list[int]:
    return sorted({y for y, _ in mmap})


# ─── Cost-center entity: seeding & department→cost-center resolution ──────────

async def _all_departments(db: AsyncSession) -> list[str]:
    rows = (await db.execute(select(distinct(Equipment.department)))).scalars().all()
    return sorted({d.strip() for d in rows if d and d.strip()}, key=str.lower)


async def _ensure_seeded(db: AsyncSession) -> None:
    """First run: turn every existing equipment department into a cost center
    (1:1), so the P&L works immediately. No-op once any cost center exists."""
    n = (await db.execute(select(func.count(CostCenter.id)))).scalar() or 0
    if n:
        return
    for i, d in enumerate(await _all_departments(db)):
        cc = CostCenter(name=d, sort_order=i)
        db.add(cc)
        await db.flush()
        db.add(CostCenterDepartment(cost_center_id=cc.id, department=d))
    await db.commit()


async def _dept_map(db: AsyncSession) -> dict[str, str]:
    """department(lower) → cost-center name (all cost centers, active or not)."""
    rows = (await db.execute(
        select(CostCenterDepartment.department, CostCenter.name)
        .join(CostCenter, CostCenterDepartment.cost_center_id == CostCenter.id)
    )).all()
    return {d.strip().lower(): name for d, name in rows if d and d.strip()}


def _resolve_cc(dept: Optional[str], mapping: dict[str, str]) -> str:
    """Cost center a department rolls up to: its mapping, else the raw department
    (an as-yet-unmapped one), else the catch-all bucket."""
    if not dept or not dept.strip():
        return UNASSIGNED
    d = dept.strip()
    return mapping.get(d.lower(), d)


def _resolve_cc_explicit(explicit: Optional[str], dept: Optional[str], mapping: dict[str, str]) -> str:
    """Cost center a cost line books to: the explicit cost center chosen at
    approval (WorkOrder/MachineIntervention.cost_center) when set, else the
    equipment department's mapping (legacy / not-yet-approved lines)."""
    e = (explicit or "").strip()
    return e if e else _resolve_cc(dept, mapping)


# ─── Site scoping (QS = Saint-Jérôme, QM = Mirabel) ───────────────────────────
#
# The plant runs from two sites. They are told apart by the cost-center name:
# Mirabel cost centers carry "Mirabel" in their name (e.g. "Maintenance -
# Mirabel", "Warehouse Mirabel"); every other cost center is Saint-Jérôme. The
# optional `site` filter narrows every aggregate on the page to one site.

SITES = ("QS", "QM")


def _site_of(cost_center: Optional[str]) -> str:
    return "QM" if cost_center and "mirabel" in cost_center.lower() else "QS"


def _site_ok(cost_center: Optional[str], site: Optional[str]) -> bool:
    """True when the cost center belongs to the requested site (or no site set)."""
    return not site or _site_of(cost_center) == site


async def _resolve_site(db: AsyncSession, ctx: PlantContext, site: Optional[str]) -> Optional[str]:
    """Constrain the QS/QM site filter to the caller's plant memberships.
    Members of both Quebec plants (and corporate) keep the free choice,
    including the combined view (None). Single-site users are locked to their
    plant's site; requesting the other one is refused. Only reached when the
    ACTIVE plant is a Quebec one — any other plant takes the strict plant_id
    scope in _resolve_scope instead."""
    codes = set((await db.execute(
        select(Plant.code).where(Plant.id.in_(list(ctx.allowed_plant_ids)))
    )).scalars().all())
    allowed = {c for c in SITES if c in codes}
    if ctx.is_corporate or len(allowed) >= 2:
        return site
    if not allowed:
        # A plant outside the QS/QM cost universe (e.g. future NL before its own
        # cost wiring) must not see the Quebec ledger at all.
        raise HTTPException(status_code=403, detail=ERR_PLANT_NOT_AUTHORIZED)
    only = next(iter(allowed))
    if site and site != only:
        raise HTTPException(status_code=403, detail=ERR_PLANT_NOT_AUTHORIZED)
    return only


async def _site_ids(db: AsyncSession) -> dict:
    """QS/QM site codes → plant ids (the Quebec cost universe). One small query
    per request; the repartition basis below prefers these over the name rule."""
    rows = (await db.execute(select(Plant.id, Plant.code).where(Plant.code.in_(SITES)))).all()
    return {code: pid for pid, code in rows}


def _row_site_ok(row_plant_id, cost_center: Optional[str], site: Optional[str], site_ids: dict) -> bool:
    """Repartition basis for one cost line: attribute by the row's plant_id when
    it has one (authoritative — set from the equipment/machine at creation), else
    fall back to the legacy cost-center NAME rule (rows predating plant stamping).
    A row owned by a plant outside QS/QM (e.g. Las Vegas) never appears on the
    Quebec cost page. `site=None` = the QS+QM combined view. This replaces the
    pure name rule (`_site_ok`); on today's data the two agree exactly (SAP lines
    100%, WO-derived tables empty), so the switch is a no-op for current numbers."""
    if row_plant_id is not None:
        qs, qm = site_ids.get("QS"), site_ids.get("QM")
        if row_plant_id not in (qs, qm):
            return False
        return site is None or row_plant_id == site_ids.get(site)
    return _site_ok(cost_center, site)


@dataclass(frozen=True)
class Scope:
    """Resolved cost scope for one request. The Quebec plants (QS/QM) share one
    cost universe — the SAP fiscal ledger, the shared budget rows — refined by
    `site` (None = the combined QS+QM view). Any OTHER active plant (e.g. Las
    Vegas) is scoped strictly by its own plant_id: only rows stamped with that
    plant count, and the Quebec ledger (including legacy NULL-plant rows, which
    all predate multi-plant and belong to Quebec) never leaks in."""
    site: Optional[str] = None
    plant_id: Optional[UUID] = None

    @property
    def is_plant(self) -> bool:
        return self.plant_id is not None


async def _resolve_scope(db: AsyncSession, ctx: PlantContext, site: Optional[str]) -> Scope:
    """Scope the page to the ACTIVE plant (X-Plant-Id), not just memberships.
    Active plant QS/QM → the Quebec universe, with the site filter constrained
    by memberships (_resolve_site, unchanged). Any other active plant → strict
    plant scoping (its own platform-tracked costs; `site` is ignored)."""
    code = (await db.execute(select(Plant.code).where(Plant.id == ctx.plant_id))).scalar_one_or_none()
    if code not in SITES:
        return Scope(plant_id=ctx.plant_id)
    return Scope(site=await _resolve_site(db, ctx, site))


def _row_scope_ok(row_plant_id, cost_center: Optional[str], scope: Scope, site_ids: dict) -> bool:
    """Whether one cost line belongs to the request's scope. Plant scope: only
    rows stamped with exactly that plant (legacy NULL-plant rows are Quebec's).
    Quebec scope: the existing plant-id-then-name-rule repartition."""
    if scope.is_plant:
        return row_plant_id == scope.plant_id
    return _row_site_ok(row_plant_id, cost_center, scope.site, site_ids)


class BudgetItem(BaseModel):
    month: int = Field(ge=1, le=12)
    amount: float = Field(ge=0)


class BudgetSave(BaseModel):
    year: int = Field(ge=2000, le=2100)
    items: List[BudgetItem]


async def _monthly_actuals(db: AsyncSession, year: int, scope: Scope) -> dict[int, float]:
    """Actual maintenance cost per month of `year`: WO cost transactions plus WO
    parts plus approved intervention parts. Scoped with the SAME repartition
    basis as _cc_actuals (_row_scope_ok: strict plant_id for a non-Quebec plant;
    plant_id-then-name-rule inside the Quebec universe)."""
    site_ids = await _site_ids(db)
    mapping = await _dept_map(db)
    totals: dict[int, float] = {m: 0.0 for m in range(1, 13)}

    def add(cc, dept, m: int, amount: float, plant_id) -> None:
        if _row_scope_ok(plant_id, _resolve_cc_explicit(cc, dept, mapping), scope, site_ids):
            totals[m] += amount

    rows = (await db.execute(
        select(WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
               extract("month", WOCost.date).label("m"), func.sum(WOCost.amount))
        .select_from(WOCost)
        .join(WorkOrder, WOCost.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOCost.date) == year)
        .group_by(WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id, extract("month", WOCost.date))
    )).all()
    for cc, dept, plant_id, m, total in rows:
        add(cc, dept, int(m), float(total or 0), plant_id)

    rows = (await db.execute(
        select(WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
               extract("month", WOPart.created_at).label("m"), func.sum(WOPart.total_cost))
        .select_from(WOPart)
        .join(WorkOrder, WOPart.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOPart.created_at) == year, WOPart.total_cost.isnot(None))
        .group_by(WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id, extract("month", WOPart.created_at))
    )).all()
    for cc, dept, plant_id, m, total in rows:
        add(cc, dept, int(m), float(total or 0), plant_id)

    rows = (await db.execute(
        select(MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
               extract("month", InterventionPart.added_at).label("m"), func.sum(InterventionPart.total_cost))
        .select_from(InterventionPart)
        .join(MachineIntervention, InterventionPart.intervention_id == MachineIntervention.id)
        .join(Equipment, MachineIntervention.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", InterventionPart.added_at) == year,
               InterventionPart.approval_status == "approved",
               InterventionPart.total_cost.isnot(None))
        .group_by(MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
                  extract("month", InterventionPart.added_at))
    )).all()
    for cc, dept, plant_id, m, total in rows:
        add(cc, dept, int(m), float(total or 0), plant_id)

    return totals


def _budget_plant_id(scope: Scope, site_ids: dict):
    """The plant a budget is stored under: a non-Quebec plant → itself; a single
    Quebec site → that plant; the combined QS+QM view → the shared row (NULL)."""
    if scope.is_plant:
        return scope.plant_id
    return site_ids.get(scope.site) if scope.site else None


async def _budgets_for(db: AsyncSession, year: int, plant_id=None, fallback_shared: bool = True) -> dict[int, float]:
    """Budget for a plant (or the shared NULL row when plant_id is None). A
    Quebec single-site view with no plant-specific budget falls back to the
    shared row; a non-Quebec plant never does (the shared row is Quebec's)."""
    cond = MaintenanceBudget.plant_id.is_(None) if plant_id is None else MaintenanceBudget.plant_id == plant_id
    rows = (await db.execute(
        select(MaintenanceBudget).where(MaintenanceBudget.year == year, cond)
    )).scalars().all()
    if not rows and plant_id is not None and fallback_shared:
        rows = (await db.execute(
            select(MaintenanceBudget).where(MaintenanceBudget.year == year, MaintenanceBudget.plant_id.is_(None))
        )).scalars().all()
    return {b.month: float(b.amount or 0) for b in rows}


@router.get("/summary")
async def cost_summary(
    year: int = Query(default=None, ge=2000, le=2100),
    site: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    ctx: PlantContext = Depends(get_plant_context),
    current_user: User = Depends(get_current_user),
):
    """Budget vs actual for one year, month by month. `ytd_*` stop at the current
    month for the current year (full year otherwise). Scoped to the ACTIVE
    plant — Quebec plants see the QS/QM universe, any other plant (e.g. Las
    Vegas) sees strictly its own plant-stamped costs and budget."""
    scope = await _resolve_scope(db, ctx, site)
    today = date.today()
    if year is None:
        year = today.year
    site_ids = await _site_ids(db)
    actuals = await _monthly_actuals(db, year, scope)
    budgets = await _budgets_for(db, year, _budget_plant_id(scope, site_ids),
                                 fallback_shared=not scope.is_plant)

    last_month = today.month if year == today.year else (12 if year < today.year else 0)
    months = [
        {"month": m, "actual": round(actuals[m], 2), "budget": round(budgets.get(m, 0.0), 2)}
        for m in range(1, 13)
    ]
    return {
        "year": year,
        "currency": "CAD",
        "months": months,
        "total_actual": round(sum(actuals.values()), 2),
        "total_budget": round(sum(budgets.get(m, 0.0) for m in range(1, 13)), 2),
        "ytd_actual": round(sum(actuals[m] for m in range(1, last_month + 1)), 2),
        "ytd_budget": round(sum(budgets.get(m, 0.0) for m in range(1, last_month + 1)), 2),
    }


@router.get("/budgets")
async def get_budgets(
    year: int = Query(..., ge=2000, le=2100),
    site: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    ctx: PlantContext = Depends(get_plant_context),
    current_user: User = Depends(get_current_user),
):
    scope = await _resolve_scope(db, ctx, site)
    site_ids = await _site_ids(db)
    budgets = await _budgets_for(db, year, _budget_plant_id(scope, site_ids),
                                 fallback_shared=not scope.is_plant)
    return [{"month": m, "amount": budgets.get(m, 0.0)} for m in range(1, 13)]


@router.put("/budgets")
async def save_budgets(
    data: BudgetSave,
    site: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    ctx: PlantContext = Depends(get_plant_context),
    current_user: User = Depends(get_current_user),
):
    """Upsert the monthly budget amounts for a year, under the caller's scope
    (non-Quebec plant → its own rows; single Quebec site → that plant; combined
    QS+QM view → the shared row)."""
    scope = await _resolve_scope(db, ctx, site)
    site_ids = await _site_ids(db)
    plant_id = _budget_plant_id(scope, site_ids)
    cond = MaintenanceBudget.plant_id.is_(None) if plant_id is None else MaintenanceBudget.plant_id == plant_id
    existing = {
        b.month: b for b in (await db.execute(
            select(MaintenanceBudget).where(MaintenanceBudget.year == data.year, cond)
        )).scalars().all()
    }
    for item in data.items:
        row = existing.get(item.month)
        if row:
            row.amount = item.amount
        else:
            db.add(MaintenanceBudget(year=data.year, month=item.month, amount=item.amount, plant_id=plant_id))
    await db.commit()
    budgets = await _budgets_for(db, data.year, plant_id, fallback_shared=not scope.is_plant)
    return [{"month": m, "amount": budgets.get(m, 0.0)} for m in range(1, 13)]


# ─── Cost centers — budgets & P&L ────────────────────────────────────────────

async def _cc_code_map(db: AsyncSession) -> dict[str, str]:
    """Cost-center SAP code (e.g. 'CA101020') keyed by name, for display."""
    return {
        name: code for name, code in (await db.execute(
            select(CostCenter.name, CostCenter.code).where(CostCenter.code.isnot(None))
        )).all() if code
    }


async def _cost_center_names(db: AsyncSession) -> list[str]:
    """Every cost center we know about: equipment departments, work-order cost
    centers, and any that already carry a budget. Sorted, blanks dropped."""
    names: set[str] = set()
    for col, model in (
        (Equipment.department, Equipment),
        (WorkOrder.cost_center, WorkOrder),
        (CostCenterBudget.cost_center, CostCenterBudget),
    ):
        rows = (await db.execute(select(distinct(col)))).scalars().all()
        for v in rows:
            if v and v.strip():
                names.add(v.strip())
    return sorted(names, key=str.lower)


# Budget-vs-Actual scoping: improvement WOs are capital projects (CAPEX); every
# other WO type is running maintenance (OPEX). Internal labor never enters the
# comparative — it's payroll, shown only in the informative by-machine view.
KINDS = ("opex", "capex")
CAPEX_WO_TYPES = {"improvement"}


def _scope_of(wo_type) -> str:
    wt = wo_type.value if hasattr(wo_type, "value") else wo_type
    return "capex" if wt in CAPEX_WO_TYPES else "opex"


async def _cc_actuals(db: AsyncSession, mmap: MonthsMap, scope: Scope) -> dict[str, dict]:
    """Per cost center and scope (opex | capex): actuals per months-map slot [12]
    and a by-expense-type breakdown, each type also [12] (so the client can
    slice by period). Costs group by the equipment department, mapped to a cost
    center; scope comes from the work-order type. Internal labor is excluded.
    Sources: WO cost transactions, WO parts, and approved intervention parts."""
    mapping = await _dept_map(db)
    site_ids = await _site_ids(db)
    slots = _slot_index(mmap)
    years = _map_years(mmap)
    out: dict[str, dict] = {}

    def add(cc: str, kind: str, y: int, m: int, amount: float, ttype: str, plant_id=None) -> None:
        i = slots.get((y, m))
        if i is None or ttype == "labor" or not _row_scope_ok(plant_id, cc, scope, site_ids):
            return
        cc_b = out.setdefault(cc, {k: {"monthly": [0.0] * 12, "by_type": {}} for k in KINDS})
        b = cc_b[kind]
        b["monthly"][i] += amount
        b["by_type"].setdefault(ttype, [0.0] * 12)[i] += amount

    # WO cost transactions, split by transaction type. Book to the WO's explicit
    # cost center (chosen at approval), falling back to the equipment department.
    rows = (await db.execute(
        select(WorkOrder.cost_center, Equipment.department, WorkOrder.type, WorkOrder.plant_id,
               extract("year", WOCost.date).label("y"), extract("month", WOCost.date).label("m"),
               WOCost.transaction_type, func.sum(WOCost.amount))
        .select_from(WOCost)
        .join(WorkOrder, WOCost.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .group_by(WorkOrder.cost_center, Equipment.department, WorkOrder.type, WorkOrder.plant_id,
                  extract("year", WOCost.date), extract("month", WOCost.date), WOCost.transaction_type)
        .where(extract("year", WOCost.date).in_(years))
    )).all()
    for cc, dept, wt, plant_id, y, m, ttype, total in rows:
        tt = ttype.value if hasattr(ttype, "value") else (ttype or "other")
        add(_resolve_cc_explicit(cc, dept, mapping), _scope_of(wt), int(y), int(m), float(total or 0), tt, plant_id)

    # WO parts used → "parts"
    rows = (await db.execute(
        select(WorkOrder.cost_center, Equipment.department, WorkOrder.type, WorkOrder.plant_id,
               extract("year", WOPart.created_at).label("y"),
               extract("month", WOPart.created_at).label("m"), func.sum(WOPart.total_cost))
        .select_from(WOPart)
        .join(WorkOrder, WOPart.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOPart.created_at).in_(years), WOPart.total_cost.isnot(None))
        .group_by(WorkOrder.cost_center, Equipment.department, WorkOrder.type, WorkOrder.plant_id,
                  extract("year", WOPart.created_at), extract("month", WOPart.created_at))
    )).all()
    for cc, dept, wt, plant_id, y, m, total in rows:
        add(_resolve_cc_explicit(cc, dept, mapping), _scope_of(wt), int(y), int(m), float(total or 0), "parts", plant_id)

    # Approved intervention parts → "parts" (kiosk interventions are reactive → OPEX)
    rows = (await db.execute(
        select(MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
               extract("year", InterventionPart.added_at).label("y"),
               extract("month", InterventionPart.added_at).label("m"),
               func.sum(InterventionPart.total_cost))
        .select_from(InterventionPart)
        .join(MachineIntervention, InterventionPart.intervention_id == MachineIntervention.id)
        .join(Equipment, MachineIntervention.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", InterventionPart.added_at).in_(years),
               InterventionPart.approval_status == "approved",
               InterventionPart.total_cost.isnot(None))
        .group_by(MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
                  extract("year", InterventionPart.added_at), extract("month", InterventionPart.added_at))
    )).all()
    for cc, dept, plant_id, y, m, total in rows:
        add(_resolve_cc_explicit(cc, dept, mapping), "opex", int(y), int(m), float(total or 0), "parts", plant_id)

    return out


async def _wo_type_actuals(db: AsyncSession, mmap: MonthsMap, scope: Scope) -> dict[str, dict[str, list]]:
    """Plant-wide actuals grouped by work-order type (corrective, preventive, …),
    each carrying a by-expense-type breakdown per months-map slot [12].
    Intervention parts count as corrective — kiosk interventions are reactive by
    nature. Internal labor is excluded (it never enters the comparative). Powers
    the planned-vs-unplanned analysis and (summed) the previous-year comparison.
    When `site` is set, each line is booked to its cost center (explicit or
    department-mapped) and dropped if it belongs to the other site."""
    mapping = await _dept_map(db)
    site_ids = await _site_ids(db)
    slots = _slot_index(mmap)
    years = _map_years(mmap)
    out: dict[str, dict[str, list]] = {}

    def add(wo_type: str, expense: str, cc: str, y: int, m: int, amount: float, plant_id=None) -> None:
        i = slots.get((y, m))
        if i is None or expense == "labor" or not _row_scope_ok(plant_id, cc, scope, site_ids):
            return
        out.setdefault(wo_type, {}).setdefault(expense, [0.0] * 12)[i] += amount

    rows = (await db.execute(
        select(WorkOrder.type, WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
               WOCost.transaction_type, extract("year", WOCost.date).label("y"),
               extract("month", WOCost.date).label("m"), func.sum(WOCost.amount))
        .select_from(WOCost)
        .join(WorkOrder, WOCost.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOCost.date).in_(years))
        .group_by(WorkOrder.type, WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
                  WOCost.transaction_type, extract("year", WOCost.date),
                  extract("month", WOCost.date))
    )).all()
    for wt, cc, dept, plant_id, ttype, y, m, total in rows:
        wt_s = wt.value if hasattr(wt, "value") else (wt or "corrective")
        tt = ttype.value if hasattr(ttype, "value") else (ttype or "other")
        add(wt_s, tt, _resolve_cc_explicit(cc, dept, mapping), int(y), int(m), float(total or 0), plant_id)

    rows = (await db.execute(
        select(WorkOrder.type, WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
               extract("year", WOPart.created_at).label("y"),
               extract("month", WOPart.created_at).label("m"), func.sum(WOPart.total_cost))
        .select_from(WOPart)
        .join(WorkOrder, WOPart.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOPart.created_at).in_(years), WOPart.total_cost.isnot(None))
        .group_by(WorkOrder.type, WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
                  extract("year", WOPart.created_at), extract("month", WOPart.created_at))
    )).all()
    for wt, cc, dept, plant_id, y, m, total in rows:
        wt_s = wt.value if hasattr(wt, "value") else (wt or "corrective")
        add(wt_s, "parts", _resolve_cc_explicit(cc, dept, mapping), int(y), int(m), float(total or 0), plant_id)

    rows = (await db.execute(
        select(MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
               extract("year", InterventionPart.added_at).label("y"),
               extract("month", InterventionPart.added_at).label("m"),
               func.sum(InterventionPart.total_cost))
        .select_from(InterventionPart)
        .join(MachineIntervention, InterventionPart.intervention_id == MachineIntervention.id)
        .join(Equipment, MachineIntervention.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", InterventionPart.added_at).in_(years),
               InterventionPart.approval_status == "approved",
               InterventionPart.total_cost.isnot(None))
        .group_by(MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
                  extract("year", InterventionPart.added_at),
                  extract("month", InterventionPart.added_at))
    )).all()
    for cc, dept, plant_id, y, m, total in rows:
        add("corrective", "parts", _resolve_cc_explicit(cc, dept, mapping), int(y), int(m), float(total or 0), plant_id)

    return out


async def _daily_actuals(db: AsyncSession, year: int, month: int, scope: Scope) -> dict[str, dict]:
    """Plant-wide actuals per day of (year, month), split by scope (opex | capex),
    each with an expense-type breakdown (also daily). Same rules as the P&L:
    internal labor excluded, scope from the WO type, intervention parts are
    corrective → OPEX. Powers the month-landing chart. When `site` is set, lines
    are booked to their cost center and the other site is dropped."""
    mapping = await _dept_map(db)
    site_ids = await _site_ids(db)
    ndays = monthrange(year, month)[1]
    out: dict[str, dict] = {k: {"daily": [0.0] * ndays, "by_type": {}} for k in KINDS}

    def add(wo_type, expense: str, cc: str, day: int, amount: float, plant_id=None) -> None:
        if expense == "labor" or not (1 <= day <= ndays) or not _row_scope_ok(plant_id, cc, scope, site_ids):
            return
        b = out[_scope_of(wo_type)]
        b["daily"][day - 1] += amount
        b["by_type"].setdefault(expense, [0.0] * ndays)[day - 1] += amount

    rows = (await db.execute(
        select(WorkOrder.type, WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
               WOCost.transaction_type,
               extract("day", WOCost.date).label("d"), func.sum(WOCost.amount))
        .select_from(WOCost)
        .join(WorkOrder, WOCost.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOCost.date) == year,
               extract("month", WOCost.date) == month)
        .group_by(WorkOrder.type, WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
                  WOCost.transaction_type, extract("day", WOCost.date))
    )).all()
    for wt, cc, dept, plant_id, ttype, d, total in rows:
        tt = ttype.value if hasattr(ttype, "value") else (ttype or "other")
        add(wt, tt, _resolve_cc_explicit(cc, dept, mapping), int(d), float(total or 0), plant_id)

    rows = (await db.execute(
        select(WorkOrder.type, WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
               extract("day", WOPart.created_at).label("d"), func.sum(WOPart.total_cost))
        .select_from(WOPart)
        .join(WorkOrder, WOPart.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOPart.created_at) == year,
               extract("month", WOPart.created_at) == month,
               WOPart.total_cost.isnot(None))
        .group_by(WorkOrder.type, WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
                  extract("day", WOPart.created_at))
    )).all()
    for wt, cc, dept, plant_id, d, total in rows:
        add(wt, "parts", _resolve_cc_explicit(cc, dept, mapping), int(d), float(total or 0), plant_id)

    rows = (await db.execute(
        select(MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
               extract("day", InterventionPart.added_at).label("d"),
               func.sum(InterventionPart.total_cost))
        .select_from(InterventionPart)
        .join(MachineIntervention, InterventionPart.intervention_id == MachineIntervention.id)
        .join(Equipment, MachineIntervention.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", InterventionPart.added_at) == year,
               extract("month", InterventionPart.added_at) == month,
               InterventionPart.approval_status == "approved",
               InterventionPart.total_cost.isnot(None))
        .group_by(MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
                  extract("day", InterventionPart.added_at))
    )).all()
    for cc, dept, plant_id, d, total in rows:
        add("corrective", "parts", _resolve_cc_explicit(cc, dept, mapping), int(d), float(total or 0), plant_id)

    return out


async def _cc_budgets_map(db: AsyncSession, mmap: MonthsMap, scope: Scope) -> dict[str, dict[str, list]]:
    """Per cost center: budget per months-map slot [12] per kind (opex | capex)."""
    site_ids = await _site_ids(db)
    slots = _slot_index(mmap)
    out: dict[str, dict[str, list]] = {}
    for b in (await db.execute(
        select(CostCenterBudget).where(CostCenterBudget.year.in_(_map_years(mmap)))
    )).scalars().all():
        i = slots.get((b.year, b.month))
        if i is None or not _row_scope_ok(b.plant_id, b.cost_center, scope, site_ids):
            continue
        kind = b.kind if b.kind in KINDS else "opex"
        arrs = out.setdefault(b.cost_center, {k: [0.0] * 12 for k in KINDS})
        arrs[kind][i] = float(b.amount or 0)
    return out


async def _sap_data(db: AsyncSession, fiscal_year: int, scope: Scope) -> Optional[dict]:
    """SAP GL lines for a fiscal year, rolled up per cost center: budget/actual
    per fiscal slot [12] plus a by-account breakdown of the actuals and the
    analyst comments. None when the fiscal year was never imported — or when the
    scope is a non-Quebec plant with no SAP lines of its own (the Quebec ledger
    never leaks; such plants stay on the internal calendar-year series)."""
    site_ids = await _site_ids(db)
    lines = (await db.execute(
        select(SapCostLine).where(SapCostLine.fiscal_year == fiscal_year)
    )).scalars().all()
    if scope.is_plant:
        lines = [ln for ln in lines if ln.plant_id == scope.plant_id]
    if not lines:
        return None
    ccs: dict[str, dict] = {}
    tot_budget = [0.0] * 12
    tot_actual = [0.0] * 12
    for ln in lines:
        if not (1 <= ln.pos <= 12) or not _row_scope_ok(ln.plant_id, ln.cost_center, scope, site_ids):
            continue
        i = ln.pos - 1
        # Display labels carry the SAP codes ("63008000 R&M - Equipment").
        acct = (f"{ln.account_code} {ln.account_name}"
                if ln.account_code and ln.account_code != ln.account_name else ln.account_name)
        cc = ccs.setdefault(ln.cost_center, {
            "code": ln.cost_center_code if ln.cost_center_code != ln.cost_center else None,
            "budget": [0.0] * 12, "actual": [0.0] * 12, "by_type": {}, "comments": [],
        })
        cc["budget"][i] += float(ln.budget or 0)
        cc["actual"][i] += float(ln.actual or 0)
        cc["by_type"].setdefault(acct, [0.0] * 12)[i] += float(ln.actual or 0)
        if ln.comment:
            cc["comments"].append({"pos": ln.pos, "account": acct, "text": ln.comment})
        tot_budget[i] += float(ln.budget or 0)
        tot_actual[i] += float(ln.actual or 0)
    for cc in ccs.values():
        cc["comments"].sort(key=lambda c: c["pos"])
    return {"cost_centers": ccs, "tot_budget": tot_budget, "tot_actual": tot_actual}


# Open purchase orders are commitments: emitted (sent/confirmed) but not yet
# received/invoiced, so not yet in the SAP actual. They land in the forecast in
# their expected-delivery month and drop out once received.
OPEN_PO_STATUSES = ("sent", "confirmed")


async def _commitments(db: AsyncSession, mmap: MonthsMap, scope: Scope) -> dict:
    """Open-PO commitments per cost center and scope: amount per months-map slot
    [12]. Placed by the PO's expected delivery month (order_date as fallback);
    only sent/confirmed POs with a cost center count. Powers the forecast."""
    site_ids = await _site_ids(db)
    slots = _slot_index(mmap)
    years = _map_years(mmap)
    per_cc: dict[str, dict] = {}
    totals = {k: [0.0] * 12 for k in KINDS}

    rows = (await db.execute(
        select(PurchaseOrder.cost_center, PurchaseOrder.scope,
               PurchaseOrder.expected_date, PurchaseOrder.order_date,
               PurchaseOrder.total_amount, PurchaseOrder.status, PurchaseOrder.plant_id)
        .where(PurchaseOrder.cost_center.isnot(None))
    )).all()
    for cc, po_scope, expected, order_date, amount, status, plant_id in rows:
        st = status.value if hasattr(status, "value") else status
        if st not in OPEN_PO_STATUSES or not cc or not _row_scope_ok(plant_id, cc, scope, site_ids):
            continue
        when = expected or order_date
        if not when:
            continue
        i = slots.get((when.year, when.month))
        if i is None:
            continue
        sc = "capex" if po_scope == "capex" else "opex"
        arrs = per_cc.setdefault(cc.strip(), {k: [0.0] * 12 for k in KINDS})
        arrs[sc][i] += float(amount or 0)
        totals[sc][i] += float(amount or 0)

    return {"per_cc": per_cc, "totals": totals}


@router.get("/cost-centers")
async def list_cost_centers(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Active cost-center names, in display order (for the budget picker)."""
    await _ensure_seeded(db)
    rows = (await db.execute(
        select(CostCenter.name).where(CostCenter.active == True)  # noqa: E712
        .order_by(CostCenter.sort_order, CostCenter.name)
    )).scalars().all()
    return list(rows)


@router.get("/pnl")
async def cost_pnl(
    year: int = Query(default=None, ge=2000, le=2100),
    site: Optional[str] = Query(default=None, pattern="^(QS|QM)$"),
    db: AsyncSession = Depends(get_db),
    ctx: PlantContext = Depends(get_plant_context),
):
    scope = await _resolve_scope(db, ctx, site)
    """Budget vs actual by cost center for a year — the cost-control statement,
    split into OPEX and CAPEX scopes (CAPEX = improvement work orders; internal
    labor excluded from both — it lives in the informative by-machine view).
    Each cost center carries per-scope budget/actual arrays [12] and a
    by-expense-type breakdown; the client slices any month range and computes
    variance. Also ships plant-wide by-WO-type actuals (planned vs unplanned
    analysis) and previous-year per-scope actuals (YoY).

    When SAP GL lines were imported for the year, `year` means the SAP FISCAL
    year (Dec–Nov): slot 0 is December of year-1, the OPEX comparative comes
    from SAP (official ledger), the by-type breakdown is per GL account, and the
    platform-tracked OPEX total ships as `tracked_actual` (coverage indicator).
    CAPEX stays platform-tracked either way. `month_map` tells the client which
    calendar (year, month) sits behind each slot."""
    if year is None:
        year = date.today().year
    await _ensure_seeded(db)
    sap = await _sap_data(db, year, scope)
    # Fiscal mode is a property of the year within the scope's universe: inside
    # Quebec the site filter narrows the rolled-up cost centers, never whether
    # the year is fiscal; a non-Quebec plant is fiscal only if it has SAP lines
    # of its own (none today → calendar mode, internal source).
    fiscal = sap is not None
    mmap = _months_map(year, fiscal)
    actuals = await _cc_actuals(db, mmap, scope)
    budgets = await _cc_budgets_map(db, mmap, scope)
    by_wo_type = await _wo_type_actuals(db, mmap, scope)
    commitments = await _commitments(db, mmap, scope)

    cc_codes = await _cc_code_map(db)

    empty_scope = {"monthly": [0.0] * 12, "by_type": {}}
    tot_budget = {k: [0.0] * 12 for k in KINDS}
    tot_actual = {k: [0.0] * 12 for k in KINDS}
    tracked_opex = [0.0] * 12
    rows_by_cc: dict[str, dict] = {}

    def blank_row(cc: str) -> dict:
        return {
            "cost_center": cc,
            "code": cc_codes.get(cc),
            "budget": {k: [0.0] * 12 for k in KINDS},
            "actual": {k: [0.0] * 12 for k in KINDS},
            "committed": {k: [0.0] * 12 for k in KINDS},
            "by_type": {k: {} for k in KINDS},
            "comments": [],
        }

    tot_committed = {k: [0.0] * 12 for k in KINDS}

    # Platform-tracked spend. With SAP data, tracked OPEX only feeds the
    # coverage indicator — the official OPEX rows come from the ledger.
    for cc in set(actuals) | set(budgets):
        b = budgets.get(cc, {k: [0.0] * 12 for k in KINDS})
        a = actuals.get(cc, {k: empty_scope for k in KINDS})
        for i in range(12):
            tracked_opex[i] += a["opex"]["monthly"][i]
        scopes = ("capex",) if fiscal else KINDS
        for k in scopes:
            if any(b[k]) or any(a[k]["monthly"]):
                row = rows_by_cc.setdefault(cc, blank_row(cc))
                row["budget"][k] = b[k]
                row["actual"][k] = a[k]["monthly"]
                row["by_type"][k] = a[k]["by_type"]
                for i in range(12):
                    tot_budget[k][i] += b[k][i]
                    tot_actual[k][i] += a[k]["monthly"][i]

    # Official SAP OPEX rows (budget, actual, by-GL-account breakdown, comments).
    if sap:
        for cc, d in sap["cost_centers"].items():
            row = rows_by_cc.setdefault(cc, blank_row(cc))
            row["budget"]["opex"] = d["budget"]
            row["actual"]["opex"] = d["actual"]
            row["by_type"]["opex"] = d["by_type"]
            row["comments"] = d["comments"]
            if d.get("code"):
                row["code"] = d["code"]
        tot_budget["opex"] = sap["tot_budget"]
        tot_actual["opex"] = sap["tot_actual"]

    # Open-PO commitments per cost center (both scopes, both calendar and SAP years).
    for cc, arrs in commitments["per_cc"].items():
        row = rows_by_cc.setdefault(cc, blank_row(cc))
        row["committed"] = arrs
    for k in KINDS:
        for i in range(12):
            tot_committed[k][i] += commitments["totals"][k][i]

    rows = []
    for row in rows_by_cc.values():
        rows.append({
            "cost_center": row["cost_center"],
            "code": row["code"],
            "budget": {k: [round(x, 2) for x in row["budget"][k]] for k in KINDS},
            "actual": {k: [round(x, 2) for x in row["actual"][k]] for k in KINDS},
            "committed": {k: [round(x, 2) for x in row["committed"][k]] for k in KINDS},
            "by_type": {k: {t: [round(x, 2) for x in arr] for t, arr in row["by_type"][k].items()}
                        for k in KINDS},
            "comments": row["comments"],
        })
    # Worst spenders first (all scopes); the catch-all bucket sinks to the bottom.
    rows.sort(key=lambda r: (r["cost_center"] == UNASSIGNED,
                             -sum(sum(r["actual"][k]) for k in KINDS)))

    # Previous-year totals per scope, for the YoY comparison — previous fiscal
    # year from SAP when imported, platform-tracked otherwise.
    prev_mmap = _months_map(year - 1, fiscal)
    prev_wo_type = await _wo_type_actuals(db, prev_mmap, scope)
    prev_actual = {k: [0.0] * 12 for k in KINDS}
    for wt, expenses in prev_wo_type.items():
        scope = _scope_of(wt)
        for arr in expenses.values():
            for i in range(12):
                prev_actual[scope][i] += arr[i]
    if fiscal:
        prev_sap = await _sap_data(db, year - 1, scope)
        prev_actual["opex"] = prev_sap["tot_actual"] if prev_sap else [0.0] * 12

    # Current-month daily actuals (only meaningful when today falls inside the
    # months map) — the client projects the month landing from the run rate.
    today = date.today()
    current_month = None
    if (today.year, today.month) in _slot_index(mmap):
        daily = await _daily_actuals(db, today.year, today.month, scope)
        # Per-type totals cut at today — the monthly by_type arrays can't be used
        # for the landing bridge because they include future-dated lines.
        # Keep net-negative types (credit memos): dropping them would make the
        # by-type MTD sum diverge from the daily series the client projects from.
        mtd_by_type = {
            k: {ty: round(sum(arr[:today.day]), 2)
                for ty, arr in daily[k]["by_type"].items()
                if round(sum(arr[:today.day]), 2) != 0}
            for k in KINDS
        }
        current_month = {
            "month": today.month,
            "today": today.day,
            "days_in_month": len(daily["opex"]["daily"]),
            "daily": {k: [round(x, 2) for x in daily[k]["daily"]] for k in KINDS},
            "mtd_by_type": mtd_by_type,
        }

    return {
        "year": year,
        "currency": "CAD",
        "fiscal": fiscal,
        "source": "sap" if sap else "internal",
        "month_map": [{"year": y, "month": m} for y, m in mmap],
        "cost_centers": rows,
        "totals": {
            "budget": {k: [round(x, 2) for x in tot_budget[k]] for k in KINDS},
            "actual": {k: [round(x, 2) for x in tot_actual[k]] for k in KINDS},
            "committed": {k: [round(x, 2) for x in tot_committed[k]] for k in KINDS},
            "tracked_actual": [round(x, 2) for x in tracked_opex] if fiscal else None,
        },
        "by_wo_type": {
            wt: {k: [round(x, 2) for x in arr] for k, arr in expenses.items()}
            for wt, expenses in by_wo_type.items()
        },
        "prev_year": year - 1,
        "prev_actual": {k: [round(x, 2) for x in prev_actual[k]] for k in KINDS},
        "current_month": current_month,
    }


async def _machine_actuals(db: AsyncSession, mmap: MonthsMap, scope: Scope) -> dict:
    """Per machine (equipment): actuals per months-map slot [12] and a
    by-expense-type breakdown. Same sources as the P&L, grouped by the work
    order's / intervention's equipment. Costs with no equipment fall under the
    'none' bucket. When `site` is set, each line is booked to its cost center
    (explicit or department-mapped) and dropped if it belongs to the other site."""
    mapping = await _dept_map(db)
    site_ids = await _site_ids(db)
    slots = _slot_index(mmap)
    years = _map_years(mmap)
    out: dict = {}

    def bucket(key, name, code):
        b = out.get(key)
        if b is None:
            b = {"equipment_id": None if key == "none" else key,
                 "name": name, "code": code, "monthly": [0.0] * 12, "by_type": {}}
            out[key] = b
        return b

    def add(eid, name, code, cc, y, m, amount, ttype, plant_id=None):
        i = slots.get((y, m))
        if i is None or not _row_scope_ok(plant_id, cc, scope, site_ids):
            return
        key = str(eid) if eid else "none"
        b = bucket(key, name, code)
        b["monthly"][i] += amount
        b["by_type"].setdefault(ttype, [0.0] * 12)[i] += amount

    # WOCost lines EXCEPT labor — internal technician labor is sourced from
    # labor_records (effective cost) below, which is the source of truth. This
    # avoids double-counting with any legacy manual WOCost 'labor' line and lets
    # the by-machine view reflect the effective-labor accounting (breaks/lunch/
    # off-shift/vacation excluded). Machine downtime / MTTR are unaffected.
    rows = (await db.execute(
        select(WorkOrder.equipment_id, Equipment.name, Equipment.code,
               WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
               extract("year", WOCost.date).label("y"), extract("month", WOCost.date).label("m"),
               WOCost.transaction_type, func.sum(WOCost.amount))
        .select_from(WOCost)
        .join(WorkOrder, WOCost.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOCost.date).in_(years),
               WOCost.transaction_type != CostTransactionType.labor)
        .group_by(WorkOrder.equipment_id, Equipment.name, Equipment.code,
                  WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
                  extract("year", WOCost.date), extract("month", WOCost.date), WOCost.transaction_type)
    )).all()
    for eid, name, code, cc, dept, plant_id, y, m, ttype, total in rows:
        tt = ttype.value if hasattr(ttype, "value") else (ttype or "other")
        add(eid, name, code, _resolve_cc_explicit(cc, dept, mapping), int(y), int(m), float(total or 0), tt, plant_id)

    # Internal effective labor cost, per machine + month, from labor_records.
    rows = (await db.execute(
        select(WorkOrder.equipment_id, Equipment.name, Equipment.code,
               WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
               extract("year", LaborRecord.date).label("y"),
               extract("month", LaborRecord.date).label("m"), func.sum(LaborRecord.labor_cost))
        .select_from(LaborRecord)
        .join(WorkOrder, LaborRecord.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", LaborRecord.date).in_(years), LaborRecord.labor_cost.isnot(None))
        .group_by(WorkOrder.equipment_id, Equipment.name, Equipment.code,
                  WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
                  extract("year", LaborRecord.date), extract("month", LaborRecord.date))
    )).all()
    for eid, name, code, cc, dept, plant_id, y, m, total in rows:
        add(eid, name, code, _resolve_cc_explicit(cc, dept, mapping), int(y), int(m), float(total or 0), "labor", plant_id)

    rows = (await db.execute(
        select(WorkOrder.equipment_id, Equipment.name, Equipment.code,
               WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
               extract("year", WOPart.created_at).label("y"),
               extract("month", WOPart.created_at).label("m"), func.sum(WOPart.total_cost))
        .select_from(WOPart)
        .join(WorkOrder, WOPart.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOPart.created_at).in_(years), WOPart.total_cost.isnot(None))
        .group_by(WorkOrder.equipment_id, Equipment.name, Equipment.code,
                  WorkOrder.cost_center, Equipment.department, WorkOrder.plant_id,
                  extract("year", WOPart.created_at), extract("month", WOPart.created_at))
    )).all()
    for eid, name, code, cc, dept, plant_id, y, m, total in rows:
        add(eid, name, code, _resolve_cc_explicit(cc, dept, mapping), int(y), int(m), float(total or 0), "parts", plant_id)

    rows = (await db.execute(
        select(MachineIntervention.equipment_id, Equipment.name, Equipment.code,
               MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
               extract("year", InterventionPart.added_at).label("y"),
               extract("month", InterventionPart.added_at).label("m"), func.sum(InterventionPart.total_cost))
        .select_from(InterventionPart)
        .join(MachineIntervention, InterventionPart.intervention_id == MachineIntervention.id)
        .join(Equipment, MachineIntervention.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", InterventionPart.added_at).in_(years),
               InterventionPart.approval_status == "approved", InterventionPart.total_cost.isnot(None))
        .group_by(MachineIntervention.equipment_id, Equipment.name, Equipment.code,
                  MachineIntervention.cost_center, Equipment.department, MachineIntervention.plant_id,
                  extract("year", InterventionPart.added_at), extract("month", InterventionPart.added_at))
    )).all()
    for eid, name, code, cc, dept, plant_id, y, m, total in rows:
        add(eid, name, code, _resolve_cc_explicit(cc, dept, mapping), int(y), int(m), float(total or 0), "parts", plant_id)

    return out


@router.get("/by-machine")
async def cost_by_machine(
    year: int = Query(default=None, ge=2000, le=2100),
    fiscal: bool = Query(default=False),
    site: Optional[str] = Query(default=None, pattern="^(QS|QM)$"),
    db: AsyncSession = Depends(get_db),
    ctx: PlantContext = Depends(get_plant_context),
):
    scope = await _resolve_scope(db, ctx, site)
    """Actual cost per machine for a year — monthly arrays + by-type breakdown.
    With `fiscal`, slots follow the SAP fiscal year (Dec of year-1 … Nov).
    The client slices the month range and sorts. Highest spenders first."""
    if year is None:
        year = date.today().year
    data = await _machine_actuals(db, _months_map(year, fiscal), scope)
    machines = [{
        "equipment_id": b["equipment_id"],
        "name": b["name"] or "—",
        "code": b["code"],
        "monthly": [round(x, 2) for x in b["monthly"]],
        "by_type": {k: [round(x, 2) for x in arr] for k, arr in b["by_type"].items()},
    } for b in data.values()]
    machines.sort(key=lambda r: -sum(r["monthly"]))
    return {"year": year, "currency": "CAD", "machines": machines}


MAX_TRANSACTION_LINES = 400


@router.get("/transactions")
async def cost_transactions(
    year: int = Query(..., ge=2000, le=2100),
    month_from: int = Query(default=1, ge=1, le=12),
    month_to: int = Query(default=12, ge=1, le=12),
    cost_center: Optional[str] = Query(default=None),
    equipment_id: Optional[UUID] = Query(default=None),
    scope: Optional[str] = Query(default=None, pattern="^(opex|capex)$"),
    site: Optional[str] = Query(default=None, pattern="^(QS|QM)$"),
    fiscal: bool = Query(default=False),
    db: AsyncSession = Depends(get_db),
    ctx: PlantContext = Depends(get_plant_context),
):
    plant_scope = await _resolve_scope(db, ctx, site)
    """Individual cost lines behind the aggregates — the audit trail. Same three
    sources as the P&L, filterable by month range, cost center, machine or site.
    `scope` mirrors the Budget-vs-Actual view: opex/capex by WO type, internal
    labor dropped; no scope = everything, labor included (by-machine view).
    With `fiscal`, month_from/month_to are fiscal slots (1 = Dec of year-1).
    Newest first, capped at MAX_TRANSACTION_LINES (count/total cover everything)."""
    mapping = await _dept_map(db)
    site_ids = await _site_ids(db)
    mmap = _months_map(year, fiscal)
    allowed = set(mmap[month_from - 1:month_to])
    years = sorted({y for y, _ in allowed})
    lines: list[dict] = []

    def keep(d, explicit, dept: Optional[str], eq_id, wo_type=None, expense: str = "", plant_id=None) -> Optional[str]:
        """Resolve the line's cost center; None means it's filtered out."""
        if (d.year, d.month) not in allowed:
            return None
        if scope and (expense == "labor" or _scope_of(wo_type) != scope):
            return None
        cc = _resolve_cc_explicit(explicit, dept, mapping)
        if cost_center and cc != cost_center:
            return None
        if not _row_scope_ok(plant_id, cc, plant_scope, site_ids):
            return None
        if equipment_id and (eq_id is None or eq_id != equipment_id):
            return None
        return cc

    rows = (await db.execute(
        select(WOCost.date, WOCost.transaction_type, WOCost.description, WOCost.amount,
               WorkOrder.id, WorkOrder.wo_number, WorkOrder.title, WorkOrder.type, WorkOrder.cost_center,
               WorkOrder.plant_id,
               Equipment.id, Equipment.name, Equipment.code, Equipment.department)
        .select_from(WOCost)
        .join(WorkOrder, WOCost.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOCost.date).in_(years))
    )).all()
    for d, ttype, desc, amount, wo_id, wo_num, wo_title, wo_type, wo_cc, wo_plant, eq_id, eq_name, eq_code, dept in rows:
        tt = ttype.value if hasattr(ttype, "value") else (ttype or "other")
        # Labor is sourced from labor_records (effective cost) below — skip any
        # legacy manual WOCost 'labor' line so the ledger doesn't double-count.
        if tt == "labor":
            continue
        cc = keep(d, wo_cc, dept, eq_id, wo_type, tt, wo_plant)
        if cc is None:
            continue
        lines.append({
            "date": d.isoformat(), "source": "wo_cost",
            "expense_type": tt,
            "wo_type": wo_type.value if hasattr(wo_type, "value") else wo_type,
            "description": desc, "amount": round(float(amount or 0), 2),
            "wo_id": str(wo_id), "wo_number": wo_num, "wo_title": wo_title,
            "equipment_name": eq_name, "equipment_code": eq_code, "cost_center": cc,
        })

    # Internal effective labor cost from labor_records (breaks/lunch/off-shift/
    # vacation already excluded). Dropped from the opex/capex scoped views by
    # keep() exactly like before; shown in the unscoped by-machine ledger.
    rows = (await db.execute(
        select(LaborRecord.date, LaborRecord.labor_cost, LaborRecord.hours_worked,
               LaborRecord.effective_hours, LaborRecord.technician_id,
               WorkOrder.id, WorkOrder.wo_number, WorkOrder.title, WorkOrder.type, WorkOrder.cost_center,
               WorkOrder.plant_id,
               Equipment.id, Equipment.name, Equipment.code, Equipment.department)
        .select_from(LaborRecord)
        .join(WorkOrder, LaborRecord.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", LaborRecord.date).in_(years), LaborRecord.labor_cost.isnot(None))
    )).all()
    for d, amount, raw_h, eff_h, tech_id, wo_id, wo_num, wo_title, wo_type, wo_cc, wo_plant, eq_id, eq_name, eq_code, dept in rows:
        cc = keep(d, wo_cc, dept, eq_id, wo_type, "labor", wo_plant)
        if cc is None:
            continue
        hrs = eff_h if eff_h is not None else raw_h
        lines.append({
            "date": d.isoformat(), "source": "labor_record", "expense_type": "labor",
            "wo_type": wo_type.value if hasattr(wo_type, "value") else wo_type,
            "description": f"{round(float(hrs or 0), 2)} h", "amount": round(float(amount or 0), 2),
            "wo_id": str(wo_id), "wo_number": wo_num, "wo_title": wo_title,
            "equipment_name": eq_name, "equipment_code": eq_code, "cost_center": cc,
        })

    rows = (await db.execute(
        select(WOPart.created_at, WOPart.description, WOPart.quantity, WOPart.total_cost,
               WorkOrder.id, WorkOrder.wo_number, WorkOrder.title, WorkOrder.type, WorkOrder.cost_center,
               WorkOrder.plant_id,
               Equipment.id, Equipment.name, Equipment.code, Equipment.department)
        .select_from(WOPart)
        .join(WorkOrder, WOPart.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", WOPart.created_at).in_(years),
               WOPart.total_cost.isnot(None))
    )).all()
    for dt, desc, qty, total, wo_id, wo_num, wo_title, wo_type, wo_cc, wo_plant, eq_id, eq_name, eq_code, dept in rows:
        cc = keep(dt, wo_cc, dept, eq_id, wo_type, "parts", wo_plant)
        if cc is None:
            continue
        qty_s = f" × {qty:g}" if qty else ""
        lines.append({
            "date": dt.date().isoformat(), "source": "wo_part", "expense_type": "parts",
            "wo_type": wo_type.value if hasattr(wo_type, "value") else wo_type,
            "description": f"{desc}{qty_s}", "amount": round(float(total or 0), 2),
            "wo_id": str(wo_id), "wo_number": wo_num, "wo_title": wo_title,
            "equipment_name": eq_name, "equipment_code": eq_code, "cost_center": cc,
        })

    rows = (await db.execute(
        select(InterventionPart.added_at, InterventionPart.item_description,
               InterventionPart.item_code, InterventionPart.quantity_used,
               InterventionPart.total_cost, MachineIntervention.intervention_type_name,
               MachineIntervention.cost_center, MachineIntervention.plant_id,
               Equipment.id, Equipment.name, Equipment.code, Equipment.department)
        .select_from(InterventionPart)
        .join(MachineIntervention, InterventionPart.intervention_id == MachineIntervention.id)
        .join(Equipment, MachineIntervention.equipment_id == Equipment.id, isouter=True)
        .where(extract("year", InterventionPart.added_at).in_(years),
               InterventionPart.approval_status == "approved",
               InterventionPart.total_cost.isnot(None))
    )).all()
    for dt, desc, code, qty, total, itype, mi_cc, mi_plant, eq_id, eq_name, eq_code, dept in rows:
        cc = keep(dt, mi_cc, dept, eq_id, "corrective", "parts", mi_plant)
        if cc is None:
            continue
        label = desc or code or itype or ""
        qty_s = f" × {qty:g}" if qty else ""
        lines.append({
            "date": dt.date().isoformat(), "source": "intervention_part", "expense_type": "parts",
            "wo_type": "corrective",
            "description": f"{label}{qty_s}", "amount": round(float(total or 0), 2),
            "wo_id": None, "wo_number": None, "wo_title": itype,
            "equipment_name": eq_name, "equipment_code": eq_code, "cost_center": cc,
        })

    lines.sort(key=lambda ln: (ln["date"], ln["amount"]), reverse=True)
    total_amount = round(sum(ln["amount"] for ln in lines), 2)
    count = len(lines)
    return {
        "year": year, "currency": "CAD", "count": count, "total_amount": total_amount,
        "truncated": count > MAX_TRANSACTION_LINES,
        "lines": lines[:MAX_TRANSACTION_LINES],
    }


# ─── Spend by supplier (procurement view) ────────────────────────────────────
#
# The official SAP ledger has no supplier dimension, so this report draws on the
# two sources that name a supplier: purchase orders (supplier + cost center +
# scope) and work-order parts (their supplier string). "received" keeps only the
# actual spend (received POs + parts used); "all" adds open commitments
# (sent/confirmed POs). Drafts and cancelled POs never count.

MAX_SUPPLIER_ORDERS = 100
OPEN_PO_SET = {"sent", "confirmed"}


@router.get("/by-supplier")
async def cost_by_supplier(
    year: int = Query(default=None, ge=2000, le=2100),
    fiscal: bool = Query(default=False),
    site: Optional[str] = Query(default=None, pattern="^(QS|QM)$"),
    status: str = Query(default="all", pattern="^(all|received)$"),
    db: AsyncSession = Depends(get_db),
    ctx: PlantContext = Depends(get_plant_context),
):
    scope = await _resolve_scope(db, ctx, site)
    """Spend per supplier for a year — the supplier expense report. Aggregates
    purchase orders (by supplier, with their cost center and OPEX/CAPEX scope)
    and work-order parts that name a supplier. `status='received'` counts only
    actual spend; `'all'` adds open-PO commitments. `site` (QS / QM) filters by
    the cost center; `fiscal` puts the window on the SAP fiscal year (Dec–Nov)."""
    if year is None:
        year = date.today().year
    mapping = await _dept_map(db)
    site_ids = await _site_ids(db)
    window = set(_months_map(year, fiscal))
    counted = {"received"} if status == "received" else {"received"} | OPEN_PO_SET

    supp: dict[str, dict] = {}

    def row(name: str) -> dict:
        s = supp.get(name)
        if s is None:
            s = {"supplier": name, "po_total": 0.0, "parts_total": 0.0,
                 "received": 0.0, "committed": 0.0, "po_count": 0,
                 "by_scope": {"opex": 0.0, "capex": 0.0}, "orders": []}
            supp[name] = s
        return s

    # Purchase orders — booked in their receipt month (expected/order date as
    # fallback). Site comes from the PO's cost center; a PO with no cost center
    # can't be placed at a site, so it's dropped when a site filter is on.
    rows = (await db.execute(
        select(PurchaseOrder.order_number, PurchaseOrder.status, PurchaseOrder.scope,
               PurchaseOrder.cost_center, PurchaseOrder.total_amount,
               PurchaseOrder.order_date, PurchaseOrder.expected_date, PurchaseOrder.received_date,
               PurchaseOrder.plant_id, Supplier.name)
        .select_from(PurchaseOrder)
        .join(Supplier, PurchaseOrder.supplier_id == Supplier.id, isouter=True)
    )).all()
    for onum, st, po_scope, cc, amount, odate, edate, rdate, po_plant, sname in rows:
        st = st.value if hasattr(st, "value") else st
        if st not in counted:
            continue
        when = rdate or edate or odate
        if not when or (when.year, when.month) not in window:
            continue
        # Stamped PO (or a plant-scoped view) → attribute by plant_id; legacy
        # no-plant PO under a Quebec site filter → original name rule, where a
        # PO without a cost center can't be placed at a site.
        if scope.is_plant or po_plant is not None:
            if not _row_scope_ok(po_plant, cc, scope, site_ids):
                continue
        elif scope.site and (not cc or _site_of(cc) != scope.site):
            continue
        amt = float(amount or 0)
        sc = "capex" if po_scope == "capex" else "opex"
        s = row((sname or "").strip() or "—")
        s["po_total"] += amt
        s["po_count"] += 1
        s["by_scope"][sc] += amt
        if st == "received":
            s["received"] += amt
        else:
            s["committed"] += amt
        if len(s["orders"]) < MAX_SUPPLIER_ORDERS:
            s["orders"].append({"order_number": onum, "date": when.isoformat(),
                                "status": st, "amount": round(amt, 2), "scope": sc,
                                "cost_center": cc})

    # Work-order parts that name a supplier — actual consumption, so kept in both
    # modes. Site resolved from the WO's cost center (or its equipment department).
    rows = (await db.execute(
        select(WOPart.supplier, WOPart.total_cost, WOPart.created_at,
               WorkOrder.cost_center, WorkOrder.type, WorkOrder.plant_id, Equipment.department)
        .select_from(WOPart)
        .join(WorkOrder, WOPart.work_order_id == WorkOrder.id)
        .join(Equipment, WorkOrder.equipment_id == Equipment.id, isouter=True)
        .where(WOPart.total_cost.isnot(None), WOPart.supplier.isnot(None))
    )).all()
    for psupplier, total, created, wcc, wtype, wo_plant, dept in rows:
        if not created or (created.year, created.month) not in window:
            continue
        cc = _resolve_cc_explicit(wcc, dept, mapping)
        if not _row_scope_ok(wo_plant, cc, scope, site_ids):
            continue
        amt = float(total or 0)
        sc = _scope_of(wtype)
        s = row((psupplier or "").strip() or "—")
        s["parts_total"] += amt
        s["by_scope"][sc] += amt

    suppliers = []
    for s in supp.values():
        total = s["po_total"] + s["parts_total"]
        if round(total, 2) == 0:
            continue
        s["orders"].sort(key=lambda o: o["date"], reverse=True)
        suppliers.append({
            "supplier": s["supplier"],
            "total": round(total, 2),
            "po_total": round(s["po_total"], 2),
            "parts_total": round(s["parts_total"], 2),
            "received": round(s["received"], 2),
            "committed": round(s["committed"], 2),
            "po_count": s["po_count"],
            "by_scope": {k: round(v, 2) for k, v in s["by_scope"].items()},
            "orders": s["orders"],
        })
    suppliers.sort(key=lambda r: -r["total"])
    return {
        "year": year, "currency": "CAD", "site": scope.site, "status": status,
        "total_amount": round(sum(r["total"] for r in suppliers), 2),
        "supplier_count": len(suppliers),
        "suppliers": suppliers,
    }


# ─── Cost-center management (CRUD + department mapping) ──────────────────────

class CostCenterCreate(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    code: Optional[str] = Field(default=None, max_length=50)


class CostCenterUpdate(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1, max_length=200)
    code: Optional[str] = Field(default=None, max_length=50)
    active: Optional[bool] = None
    sort_order: Optional[int] = None


class DeptMapItem(BaseModel):
    department: str = Field(min_length=1, max_length=200)
    cost_center_id: Optional[UUID] = None


class DeptMapSave(BaseModel):
    items: List[DeptMapItem]


@router.get("/cost-centers/manage")
async def manage_cost_centers(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Full cost-center list with their mapped departments, plus every department
    available to assign — powers the management screen."""
    await _ensure_seeded(db)
    ccs = (await db.execute(
        select(CostCenter).order_by(CostCenter.sort_order, CostCenter.name)
    )).scalars().all()
    maps = (await db.execute(select(CostCenterDepartment))).scalars().all()
    by_cc: dict[str, list] = {}
    for m in maps:
        by_cc.setdefault(str(m.cost_center_id), []).append(m.department)
    return {
        "cost_centers": [{
            "id": str(c.id), "name": c.name, "code": c.code, "active": c.active,
            "sort_order": c.sort_order,
            "departments": sorted(by_cc.get(str(c.id), []), key=str.lower),
        } for c in ccs],
        "departments": await _all_departments(db),
    }


@router.post("/cost-centers")
async def create_cost_center(
    data: CostCenterCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    name = data.name.strip()
    dup = (await db.execute(
        select(CostCenter).where(func.lower(CostCenter.name) == name.lower())
    )).scalars().first()
    if dup:
        raise HTTPException(status_code=400, detail="cost_center_exists")
    top = (await db.execute(select(func.max(CostCenter.sort_order)))).scalar() or 0
    cc = CostCenter(name=name, code=(data.code or None), sort_order=top + 1)
    db.add(cc)
    await db.commit()
    return {"id": str(cc.id), "name": cc.name, "code": cc.code, "active": cc.active, "sort_order": cc.sort_order}


@router.patch("/cost-centers/{cc_id}")
async def update_cost_center(
    cc_id: UUID,
    data: CostCenterUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cc = await db.get(CostCenter, cc_id)
    if not cc:
        raise HTTPException(status_code=404, detail="cost_center_not_found")
    if data.name and data.name.strip() != cc.name:
        new = data.name.strip()
        dup = (await db.execute(
            select(CostCenter).where(func.lower(CostCenter.name) == new.lower(), CostCenter.id != cc_id)
        )).scalars().first()
        if dup:
            raise HTTPException(status_code=400, detail="cost_center_exists")
        # keep budgets attached to the renamed cost center
        await db.execute(sa_update(CostCenterBudget)
                         .where(CostCenterBudget.cost_center == cc.name).values(cost_center=new))
        cc.name = new
    if data.code is not None:
        cc.code = data.code or None
    if data.active is not None:
        cc.active = data.active
    if data.sort_order is not None:
        cc.sort_order = data.sort_order
    await db.commit()
    return {"id": str(cc.id), "name": cc.name, "code": cc.code, "active": cc.active, "sort_order": cc.sort_order}


@router.delete("/cost-centers/{cc_id}")
async def delete_cost_center(
    cc_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cc = await db.get(CostCenter, cc_id)
    if not cc:
        raise HTTPException(status_code=404, detail="cost_center_not_found")
    await db.execute(sa_delete(CostCenterDepartment).where(CostCenterDepartment.cost_center_id == cc_id))
    await db.execute(sa_delete(CostCenterBudget).where(CostCenterBudget.cost_center == cc.name))
    await db.delete(cc)
    await db.commit()
    return {"ok": True}


@router.put("/cost-center-departments")
async def save_dept_map(
    data: DeptMapSave,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Assign departments to cost centers. Each item re-homes one department
    (cost_center_id=null unassigns it → rolls up to 'Unassigned')."""
    for item in data.items:
        dept = item.department.strip()
        if not dept:
            continue
        await db.execute(sa_delete(CostCenterDepartment)
                         .where(func.lower(CostCenterDepartment.department) == dept.lower()))
        if item.cost_center_id:
            db.add(CostCenterDepartment(cost_center_id=item.cost_center_id, department=dept))
    await db.commit()
    return {"ok": True}


class CCBudgetItem(BaseModel):
    cost_center: str = Field(min_length=1, max_length=200)
    month: int = Field(ge=1, le=12)
    amount: float = Field(ge=0)


class CCBudgetSave(BaseModel):
    year: int = Field(ge=2000, le=2100)
    kind: str = Field(default="opex", pattern="^(opex|capex)$")
    items: List[CCBudgetItem]


def _cc_budget_rows(budgets: dict[str, dict[str, list]], kind: str,
                    codes: dict[str, str] | None = None) -> list:
    codes = codes or {}
    return [{"cost_center": cc, "code": codes.get(cc), "months": [round(x, 2) for x in arrs[kind]]}
            for cc, arrs in sorted(budgets.items(), key=lambda kv: kv[0].lower())]


@router.get("/cost-center-budgets")
async def get_cc_budgets(
    year: int = Query(..., ge=2000, le=2100),
    kind: str = Query(default="opex", pattern="^(opex|capex)$"),
    site: Optional[str] = Query(default=None, pattern="^(QS|QM)$"),
    db: AsyncSession = Depends(get_db),
    ctx: PlantContext = Depends(get_plant_context),
):
    scope = await _resolve_scope(db, ctx, site)
    """Editable monthly budget grid per cost center. When the year was imported
    from SAP, the OPEX budget is the SAP ledger — return it read-only (fiscal
    Dec–Nov, with codes) instead of the manual grid, so it matches the Budget-vs-
    Actual tab. CAPEX and non-SAP years stay editable manual budgets. `site`
    narrows the grid to one plant (QS / QM)."""
    codes = await _cc_code_map(db)
    if kind == "opex":
        sap = await _sap_data(db, year, scope)
        if sap:
            mmap = _months_map(year, True)
            rows = [{
                "cost_center": cc,
                "code": d.get("code") or codes.get(cc),
                "months": [round(x, 2) for x in d["budget"]],
            } for cc, d in sorted(sap["cost_centers"].items(), key=lambda kv: kv[0].lower())]
            return {"rows": rows, "read_only": True, "source": "sap",
                    "month_map": [{"year": y, "month": m} for y, m in mmap]}
    rows = _cc_budget_rows(await _cc_budgets_map(db, _months_map(year, False), scope), kind, codes)
    return {"rows": rows, "read_only": False, "source": "internal", "month_map": None}


@router.put("/cost-center-budgets")
async def save_cc_budgets(
    data: CCBudgetSave,
    db: AsyncSession = Depends(get_db),
    ctx: PlantContext = Depends(get_plant_context),
    current_user: User = Depends(get_current_user),
):
    """Upsert monthly budget amounts per cost center for a year, per kind
    (opex | capex envelope). OPEX of a SAP-imported year is read-only (it comes
    from the ledger) — reject the write. A non-Quebec plant (e.g. Las Vegas)
    keeps its own rows, stamped with its plant_id; Quebec rows stay on the
    legacy NULL-plant + cost-center-name repartition."""
    scope = await _resolve_scope(db, ctx, None)
    if data.kind == "opex" and await _sap_data(db, data.year, scope):
        raise HTTPException(status_code=400, detail="opex_budget_from_sap")
    plant_cond = (CostCenterBudget.plant_id == scope.plant_id) if scope.is_plant \
        else CostCenterBudget.plant_id.is_(None)
    existing = {
        (b.cost_center, b.month): b
        for b in (await db.execute(
            select(CostCenterBudget).where(CostCenterBudget.year == data.year,
                                           CostCenterBudget.kind == data.kind,
                                           plant_cond)
        )).scalars().all()
    }
    for item in data.items:
        cc = item.cost_center.strip()
        if not cc:
            continue
        row = existing.get((cc, item.month))
        if row:
            row.amount = item.amount
        else:
            db.add(CostCenterBudget(year=data.year, month=item.month, kind=data.kind,
                                    cost_center=cc, amount=item.amount, plant_id=scope.plant_id))
    await db.commit()
    rows = _cc_budget_rows(await _cc_budgets_map(db, _months_map(data.year, False), scope), data.kind,
                           await _cc_code_map(db))
    return {"rows": rows, "read_only": False, "source": "internal", "month_map": None}


# ─── SAP GL import (monthly SAC extract, one workbook per fiscal year) ────────
#
# The extract carries one sheet per fiscal month (Dec–Nov); every sheet is a
# cumulative snapshot with "<Mon> (<FY>)-Budget (YTD)" / "Actuals (YTD)" columns
# per (cost center, GL account). Monthly values are recomputed here as YTD
# deltas between consecutive sheets — more robust than the sheet's own monthly
# columns and it back-applies reversals posted in later months. A re-import
# replaces the whole fiscal year.

_SAP_HDR_RE = re.compile(r"^([A-Za-z]{3})\w*\s*\((\d{4})\)")
_SAP_MONTHS = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
               "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def _fiscal_pos(month: int) -> int:
    """Calendar month → fiscal slot (Dec = 1 … Nov = 12)."""
    return (month % 12) + 1


def _split_code(raw: str) -> tuple:
    """'CA101020 Maintenance' → ('CA101020', 'Maintenance'); no code → ('', raw)."""
    s = (raw or "").strip()
    code, _, rest = s.partition(" ")
    if rest and re.fullmatch(r"[A-Za-z]{0,4}\d{3,}", code):
        return code, rest.strip()
    return "", s


def _parse_sap_workbook(content: bytes) -> dict:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue
        hdrs = [str(h).strip() if h is not None else "" for h in headers]
        bud_col = act_col = com_col = None
        fy = month = None
        for i, h in enumerate(hdrs):
            if h.endswith("Budget (YTD)"):
                bud_col = i
            elif h.endswith("Actuals (YTD)"):
                act_col = i
            elif "comment" in h.lower():
                com_col = i
            m = _SAP_HDR_RE.match(h)
            if m and fy is None and m.group(1).lower() in _SAP_MONTHS:
                month = _SAP_MONTHS[m.group(1).lower()]
                fy = int(m.group(2))
        if bud_col is None or act_col is None or fy is None:
            continue  # not a SAP budget sheet — skip quietly
        entries: dict = {}
        for r in rows:
            cc_raw = str(r[0]).strip() if r and r[0] else ""
            acct_raw = str(r[1]).strip() if r and len(r) > 1 and r[1] else ""
            if not cc_raw or not acct_raw:
                continue

            def num(i):
                v = r[i] if i is not None and i < len(r) else None
                return float(v) if isinstance(v, (int, float)) else 0.0

            e = entries.setdefault((cc_raw, acct_raw), [0.0, 0.0, None])
            e[0] += num(bud_col)
            e[1] += num(act_col)
            if com_col is not None and com_col < len(r) and r[com_col]:
                c = str(r[com_col]).strip()
                e[2] = f"{e[2]}\n{c}" if e[2] else c
        sheets.append({"fy": fy, "pos": _fiscal_pos(month),
                       "year": fy - 1 if month == 12 else fy, "month": month,
                       "entries": entries})
    if not sheets:
        raise HTTPException(status_code=400, detail="sap_import_bad_format")
    if len({s["fy"] for s in sheets}) > 1:
        raise HTTPException(status_code=400, detail="sap_import_mixed_years")
    sheets.sort(key=lambda s: s["pos"])
    if len({s["pos"] for s in sheets}) != len(sheets):
        raise HTTPException(status_code=400, detail="sap_import_duplicate_months")
    return {"fiscal_year": sheets[0]["fy"], "sheets": sheets}


async def _ensure_sap_cost_centers(db: AsyncSession, ccs: dict) -> None:
    """Make every SAP cost center a manageable entity (matched by SAP code, then
    by name) so departments can be mapped to it on the Manage tab."""
    existing = (await db.execute(select(CostCenter))).scalars().all()
    codes = {c.code for c in existing if c.code}
    names = {c.name.strip().lower() for c in existing}
    top = max((c.sort_order for c in existing), default=0)
    for code, name in sorted(ccs.items()):
        if (code and code in codes) or name.strip().lower() in names:
            continue
        top += 1
        db.add(CostCenter(name=name, code=code or None, sort_order=top))
        names.add(name.strip().lower())


@router.post("/sap-import")
async def import_sap_costs(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import the monthly SAP/SAC budget extract (.xlsx). Replaces every line of
    the workbook's fiscal year; the Costs page then treats that year as a fiscal
    year (Dec–Nov) with SAP as the official OPEX Budget-vs-Actual series."""
    content = await file.read()
    try:
        parsed = _parse_sap_workbook(content)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="sap_import_bad_format")

    fy = parsed["fiscal_year"]
    prev: dict = {}
    merged: dict = {}   # (pos, cc_code, acct_code) → SapCostLine fields
    for s in parsed["sheets"]:
        for key, (bud, act, com) in s["entries"].items():
            pb, pa = prev.get(key, (0.0, 0.0))
            d_bud, d_act = bud - pb, act - pa
            prev[key] = (bud, act)
            if abs(d_bud) < 0.005 and abs(d_act) < 0.005 and not com:
                continue
            cc_code, cc_name = _split_code(key[0])
            acct_code, acct_name = _split_code(key[1])
            mk = (s["pos"], cc_code or cc_name, acct_code or acct_name)
            row = merged.setdefault(mk, {
                "fiscal_year": fy, "pos": s["pos"], "year": s["year"], "month": s["month"],
                "cost_center_code": cc_code or cc_name, "cost_center": cc_name or cc_code,
                "account_code": acct_code or acct_name, "account_name": acct_name or acct_code,
                "budget": 0.0, "actual": 0.0, "comment": None,
            })
            row["budget"] = round(row["budget"] + d_bud, 2)
            row["actual"] = round(row["actual"] + d_act, 2)
            if com:
                row["comment"] = f"{row['comment']}\n{com}" if row["comment"] else com

    await db.execute(sa_delete(SapCostLine).where(SapCostLine.fiscal_year == fy))
    for fields in merged.values():
        db.add(SapCostLine(**fields))

    sap_ccs = {}
    for fields in merged.values():
        sap_ccs.setdefault(fields["cost_center_code"], fields["cost_center"])
    await _ensure_sap_cost_centers(db, sap_ccs)
    await db.commit()

    return {
        "fiscal_year": fy,
        "months": len(parsed["sheets"]),
        "lines": len(merged),
        "cost_centers": len(sap_ccs),
        "total_budget": round(sum(f["budget"] for f in merged.values()), 2),
        "total_actual": round(sum(f["actual"] for f in merged.values()), 2),
        "currency": "CAD",
    }
